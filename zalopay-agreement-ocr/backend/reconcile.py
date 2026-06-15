"""
Reconcile Excel internal fee data vs OCR'd agreement (fee assignments).
"""
from __future__ import annotations

import re
import unicodedata
from io import BytesIO
from typing import Any

import openpyxl

# ── Excel column positions (0-based) ─────────────────────────────────────────
FEE_PCT_COL: dict[str, int] = {
    "Ví/Balance":    12,   # M
    "Ví/ATM":        16,   # Q
    "Ví/TK NH":      20,   # U
    "Ví/VietQR":     24,   # Y
    "Ví/TKTS":       28,   # AC
    "Ví/Số dư SL":   32,   # AG
    "Ví/CC TN":      36,   # AK
    "Ví/CC NN":      40,   # AO
    "Cổng/ATM":      44,   # AS
    "Cổng/TK NH":    48,   # AW
    "Cổng/VietQR":   52,   # BA
    "Cổng/TKTS":     56,   # BE
    "Cổng/Số dư SL": 60,   # BI
    "Cổng/CC TN":    64,   # BM
    "Cổng/CC NN":    68,   # BQ
}

REFUND_FEE_COL: dict[str, int] = {
    "Ví/Balance":    71,   # BT
    "Ví/ATM":        77,   # BZ
    "Ví/TK NH":      83,   # CF
    "Ví/VietQR":     89,   # CL
    "Ví/TKTS":       95,   # CR
    "Ví/Số dư SL":   101,  # CX
    "Ví/CC TN":      107,  # DD
    "Ví/CC NN":      113,  # DJ
    "Cổng/ATM":      119,  # DP
    "Cổng/TK NH":    125,  # DV
    "Cổng/VietQR":   131,  # EB
    "Cổng/TKTS":     137,  # EH
    "Cổng/Số dư SL": 143,  # EN
    "Cổng/CC TN":    149,  # ET
    "Cổng/CC NN":    155,  # EZ
}

COL_MA_AGREE     = 1   # B
COL_TEN_AGREE    = 2   # C
COL_TRANG_THAI   = 3   # D
COL_TEN_MERCHANT = 4   # E
COL_THUE_VAT     = 9   # J — YES = chưa gồm VAT, NO = đã gồm VAT

# ── Internal channel → fee basket ────────────────────────────────────────────
# Quy tắc đặc biệt: Cổng/TKTS, Cổng/Số dư SL → rổ Kênh Zalopay App
KENH_TO_RO: dict[str, str] = {
    "Ví/Balance":    "zalopay_app",
    "Ví/ATM":        "zalopay_app",
    "Ví/TK NH":      "zalopay_app",
    "Ví/VietQR":     "zalopay_app",
    "Ví/TKTS":       "zalopay_app",
    "Ví/Số dư SL":   "zalopay_app",
    "Ví/CC TN":      "zalopay_app",
    "Ví/CC NN":      "zalopay_app",
    "Cổng/ATM":      "the_noi_dia",
    "Cổng/TK NH":    "the_noi_dia",
    "Cổng/VietQR":   "qr_da_nang",
    "Cổng/TKTS":     "zalopay_app",   # đặc biệt
    "Cổng/Số dư SL": "zalopay_app",   # đặc biệt
    "Cổng/CC TN":    "the_quoc_te",
    "Cổng/CC NN":    "the_quoc_te",
}

SPECIAL_CHANNELS = {"Cổng/TKTS", "Cổng/Số dư SL"}

# ── T1 column label → basket key (must match FeeWorkbench T1 cols exactly) ───
COL_LABEL_TO_RO: dict[str, str] = {
    "Kênh Zalopay App":                              "zalopay_app",
    "Kênh khác > Zalopay Gateway > Thẻ nội địa":    "the_noi_dia",
    "Kênh khác > Zalopay Gateway > Thẻ quốc tế":    "the_quoc_te",
    "Kênh khác > Quét Mã QR đa năng":               "qr_da_nang",
}

ROW_PHI_GD   = "Phí dịch vụ"
ROW_PHI_HOAN = "Phí xử lý hoàn trả"

VAT_RATE      = 0.10
SAI_SO        = 0.01   # điểm phần trăm được phép chênh
NGUONG_DUYET  = 90
NGUONG_REVIEW = 70


def _vi_norm(s: str) -> str:
    """Lowercase, strip diacritics — for fuzzy merchant name matching."""
    s = unicodedata.normalize("NFD", s.lower().strip())
    return "".join(c for c in s if unicodedata.category(c) != "Mn")


# ── Parsers ───────────────────────────────────────────────────────────────────

def _parse_pct(s: Any) -> float | None:
    """'1%', '1,1%/giao dịch', '0.85%' → float (percent unit, not decimal)."""
    if s is None:
        return None
    if isinstance(s, (int, float)):
        return float(s)
    s = str(s).strip()
    if not s or s in ("-", "—"):
        return None
    m = re.search(r"[\d]+[.,]?[\d]*", s)
    if not m:
        return None
    try:
        return float(m.group().replace(",", "."))
    except ValueError:
        return None


def _fmt_pct(val: float) -> str:
    """Normalize percentage display: 1.0 → '1%', 0.85 → '0.85%', 2.29 → '2.29%'."""
    return f"{val:g}%"


def _parse_vnd(s: Any) -> float | None:
    """'Miễn phí'→0, '1.000'→1000, '1000 VND'→1000."""
    if s is None:
        return None
    if isinstance(s, (int, float)):
        return float(s)
    s = str(s).strip()
    if s.lower() in ("miễn phí", "free", "0", ""):
        return 0.0
    cleaned = re.sub(r"[^\d]", "", s)
    return float(cleaned) if cleaned else None


def _excel_pct(val: Any) -> float | None:
    """
    Excel percentage cells: openpyxl returns 0.01 for cells formatted as 1%.
    But some sheets store raw numbers (1.0 for 1%).
    Heuristic: ZaloPay fees are 0.4–2.5% — values ≤ 0.3 are decimal-encoded.
    """
    if val is None or val == "":
        return None
    try:
        f = float(val)
        if f == 0:
            return None  # treat 0 as empty (kênh chưa mở)
        if 0 < f <= 0.3:
            return round(f * 100, 6)
        return f
    except (TypeError, ValueError):
        return None


def _excel_vnd(val: Any) -> float | None:
    if val is None or val == "":
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


# ── Contract fee extraction ───────────────────────────────────────────────────

def _extract_contract_baskets(fee_data: dict) -> dict[str, dict]:
    """
    Pull fee values from the FeeWorkbench assignments dict.
    Returns {basket_key: {phi_gd: float|None, phi_hoan: float|None, raw_*: str}}
    """
    assignments: dict = (fee_data or {}).get("assignments") or {}
    baskets: dict = {ro: {"phi_gd": None, "phi_hoan": None, "raw_phi_gd": None, "raw_phi_hoan": None}
                     for ro in COL_LABEL_TO_RO.values()}

    for col_label, ro_key in COL_LABEL_TO_RO.items():
        gd_key    = f"{ROW_PHI_GD}|||{col_label}"
        hoan_key  = f"{ROW_PHI_HOAN}|||{col_label}"

        if gd_key in assignments:
            raw = assignments[gd_key].get("raw", "")
            baskets[ro_key]["raw_phi_gd"] = raw
            baskets[ro_key]["phi_gd"] = _parse_pct(raw)

        if hoan_key in assignments:
            raw = assignments[hoan_key].get("raw", "")
            baskets[ro_key]["raw_phi_hoan"] = raw
            baskets[ro_key]["phi_hoan"] = _parse_vnd(raw)

    return baskets


# ── Merchant row finder ───────────────────────────────────────────────────────

def _find_merchant_row(
    ws: Any,
    ten_merchant: str | None,
    ma_so_thue: str | None,
) -> tuple[tuple | None, list[str]]:
    warnings: list[str] = []
    matches: list[tuple] = []

    ten_norm  = _vi_norm(ten_merchant or "")
    mst_lower = (ma_so_thue or "").lower().strip()

    for row in ws.iter_rows(min_row=5, values_only=True):
        if not any(row):
            continue
        row_ten  = _vi_norm(str(row[COL_TEN_MERCHANT] or ""))
        row_mst  = str(row[COL_MA_AGREE] or "").lower().strip()
        row_name_col2 = _vi_norm(str(row[COL_TEN_AGREE] or ""))

        hit = False
        if mst_lower and (mst_lower in row_ten or mst_lower in row_mst):
            hit = True
        if ten_norm and len(ten_norm) >= 4:
            # fuzzy: first 8 normalized chars overlap (handles diacritics & spacing)
            if ten_norm[:8] in row_ten or row_ten[:8] in ten_norm:
                hit = True
            if ten_norm[:8] in row_name_col2 or row_name_col2[:8] in ten_norm:
                hit = True
        if hit:
            matches.append(row)

    if len(matches) == 0:
        warnings.append(f"Không tìm thấy merchant '{ten_merchant}' (MST: {ma_so_thue}) trong file Excel.")
        return None, warnings
    if len(matches) > 1:
        names = [str(r[COL_TEN_MERCHANT]) for r in matches]
        warnings.append(f"Tìm thấy {len(matches)} dòng khớp: {names}. Không thể tự chọn — cần kiểm tra.")
        return None, warnings

    return matches[0], warnings


# ── Main reconcile function ───────────────────────────────────────────────────

def reconcile(excel_bytes: bytes, agreement_data: dict) -> dict:
    """
    Compare Excel row vs agreement OCR data.
    agreement_data = full ExtractedData dict (partner + payment + fee).
    """
    partner   = agreement_data.get("partner") or {}
    fee_data  = agreement_data.get("fee") or {}

    ten_merchant = partner.get("ten_doi_tac")
    ma_so_thue   = partner.get("ma_so_thue")
    ma_agreement = partner.get("ma_agreement")

    canh_bao: list[str] = []
    field_lech: list[dict] = []

    # ── Load workbook ────────────────────────────────────────────────────────
    try:
        wb = openpyxl.load_workbook(BytesIO(excel_bytes))
    except Exception as e:
        return {"error": f"Không đọc được file Excel: {e}"}

    if "Phí thường" not in wb.sheetnames:
        return {"error": "Không tìm thấy sheet 'Phí thường' trong file Excel."}

    ws = wb["Phí thường"]

    # ── Find merchant row ────────────────────────────────────────────────────
    row, row_warnings = _find_merchant_row(ws, ten_merchant, ma_so_thue)
    canh_bao.extend(row_warnings)

    if row is None:
        return {
            "merchant": {"ma_agreement": ma_agreement, "ten_merchant": ten_merchant,
                         "ma_so_thue": ma_so_thue, "trang_thai": ""},
            "canh_bao": canh_bao,
            "error": canh_bao[0] if canh_bao else "Không tìm thấy merchant.",
        }

    # ── Read VAT basis from column J ─────────────────────────────────────────
    vat_j_raw = str(row[COL_THUE_VAT] or "").strip().upper()
    if vat_j_raw == "YES":
        excel_vat_included = False   # J=YES → phí Excel chưa gồm VAT
        excel_co_so_vat = "chưa gồm VAT (J=YES)"
    elif vat_j_raw == "NO":
        excel_vat_included = True    # J=NO  → phí Excel đã gồm VAT
        excel_co_so_vat = "đã gồm VAT (J=NO)"
    else:
        excel_vat_included = False   # mặc định: chưa gồm VAT
        excel_co_so_vat = "VAT chưa rõ (mặc định: chưa gồm)"
        canh_bao.append(
            f"Cột J (Thuế VAT) {'trống' if not vat_j_raw else f'giá trị lạ: {vat_j_raw!r}'}. "
            "Mặc định: chưa gồm VAT."
        )

    # ── Check assignments exist ──────────────────────────────────────────────
    assignments = (fee_data or {}).get("assignments") or {}
    if not assignments:
        canh_bao.append("Biểu phí HĐ chưa được gán (FeeWorkbench trống). Kết quả đối soát sẽ không chính xác.")

    # ── Extract contract baskets ─────────────────────────────────────────────
    baskets = _extract_contract_baskets(fee_data)
    bao_gom_vat: bool | None = fee_data.get("vat_included")

    # ── VAT normalization: đưa cả 2 phía về pre-VAT trước khi so ────────────
    da_chuan_hoa_hd = False
    if bao_gom_vat is True:
        for ro in baskets.values():
            if ro["phi_gd"] is not None:
                ro["phi_gd"] = round(ro["phi_gd"] / (1 + VAT_RATE), 6)
        da_chuan_hoa_hd = True

    # ── Per-channel comparison ───────────────────────────────────────────────
    tu_khoa_excel: list[dict] = []
    tong_field = 0
    so_khop = 0

    for kenh in FEE_PCT_COL:
        excel_pct_raw = _excel_pct(row[FEE_PCT_COL[kenh]])
        # Normalize Excel phí GD to pre-VAT if J=NO (đã gồm VAT)
        if excel_vat_included and excel_pct_raw is not None:
            excel_pct = round(excel_pct_raw / (1 + VAT_RATE), 6)
        else:
            excel_pct = excel_pct_raw
        excel_hoan = _excel_vnd(row[REFUND_FEE_COL[kenh]]) if excel_pct_raw is not None else None
        kenh_mo    = excel_pct_raw is not None
        ro_key     = KENH_TO_RO[kenh]
        basket     = baskets.get(ro_key, {})
        ky_vong_pct  = basket.get("phi_gd")
        ky_vong_hoan = basket.get("phi_hoan")
        is_special   = kenh in SPECIAL_CHANNELS

        pct_str  = _fmt_pct(excel_pct) if excel_pct is not None else ""
        hoan_str = ("Miễn phí" if excel_hoan == 0 else
                    f"{int(excel_hoan):,} VND" if excel_hoan is not None else "")
        ky_vong_pct_str = _fmt_pct(ky_vong_pct) if ky_vong_pct is not None else ""

        item: dict = {
            "kenh": kenh,
            "co_mo": kenh_mo,
            "phi_gd": pct_str,
            "phi_hoan": hoan_str,
            "ro_phi": ro_key,
            "la_kenh_dac_biet": is_special,
            "ky_vong_pct_str": ky_vong_pct_str,
            "khop_pct": False,
        }

        if not kenh_mo:
            tu_khoa_excel.append(item)
            if ky_vong_pct is not None:
                canh_bao.append(f"{kenh}: HĐ có rổ '{ro_key}' nhưng Excel trống (kênh chưa mở).")
            continue

        # phí GD — so sánh số, bỏ qua text
        tong_field += 1
        if ky_vong_pct is None:
            field_lech.append({
                "kenh": kenh,
                "loai_field": "phí GD (%)",
                "gia_tri_excel": pct_str,
                "gia_tri_ky_vong": "Không áp dụng",
                "chenh_lech": "kênh mở nhưng HĐ không có rổ",
                "ghi_chu": ("kênh đặc biệt → rổ App" if is_special else
                            f"rổ '{ro_key}' không có trong HĐ"),
            })
        elif abs(excel_pct - ky_vong_pct) <= SAI_SO:
            so_khop += 1
            item["khop_pct"] = True
        else:
            chenh = round(excel_pct - ky_vong_pct, 4)
            field_lech.append({
                "kenh": kenh,
                "loai_field": "phí GD (%)",
                "gia_tri_excel": pct_str,
                "gia_tri_ky_vong": ky_vong_pct_str,
                "chenh_lech": f"{'+' if chenh > 0 else ''}{chenh}pp",
                "ghi_chu": "kênh đặc biệt (áp dụng mức App)" if is_special else "",
            })

        tu_khoa_excel.append(item)

        # phí hoàn trả — so sánh số
        if excel_hoan is not None:
            tong_field += 1
            if ky_vong_hoan is None:
                canh_bao.append(f"{kenh}: phí hoàn trả không có trong HĐ — bỏ qua.")
                so_khop += 1
            elif abs(excel_hoan - ky_vong_hoan) < 1:  # numeric comparison, 1 VND tolerance
                so_khop += 1
            else:
                kv_str = ("Miễn phí" if ky_vong_hoan == 0 else f"{int(ky_vong_hoan):,} VND")
                field_lech.append({
                    "kenh": kenh,
                    "loai_field": "phí hoàn trả",
                    "gia_tri_excel": hoan_str,
                    "gia_tri_ky_vong": kv_str,
                    "chenh_lech": f"{excel_hoan - ky_vong_hoan:+,.0f} VND",
                    "ghi_chu": "",
                })

    # ── Score ────────────────────────────────────────────────────────────────
    ty_le = round(so_khop / tong_field * 100, 1) if tong_field > 0 else 0.0
    if ty_le >= NGUONG_DUYET:
        de_xuat = "DUYỆT"
    elif ty_le >= NGUONG_REVIEW:
        de_xuat = "REVIEW"
    else:
        de_xuat = "TỪ CHỐI"

    return {
        "merchant": {
            "ma_agreement": str(row[COL_MA_AGREE] or ""),
            "ten_merchant": str(row[COL_TEN_MERCHANT] or ""),
            "ma_so_thue": ma_so_thue,
            "trang_thai": str(row[COL_TRANG_THAI] or ""),
        },
        "vat": {
            "hop_dong_bao_gom_vat": bao_gom_vat,
            "excel_co_j": vat_j_raw or "(trống)",
            "excel_co_so_vat": excel_co_so_vat,
            "da_chuan_hoa_hop_dong": da_chuan_hoa_hd,
            "da_chuan_hoa_excel": excel_vat_included,
        },
        "tu_khoa_excel": tu_khoa_excel,
        "so_sanh": {
            "tong_field_so_sanh": tong_field,
            "so_field_khop": so_khop,
            "ty_le_khop_phan_tram": ty_le,
        },
        "field_lech": field_lech,
        "canh_bao": canh_bao,
        "de_xuat": de_xuat,
        "ly_do_de_xuat": (
            f"Tỷ lệ khớp {ty_le}% ({so_khop}/{tong_field} field). "
            f"{len(field_lech)} field lệch."
        ),
        "luu_y": "Đây chỉ là gợi ý; quyết định duyệt cuối cùng do con người.",
    }
